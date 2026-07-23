"""
Le os PDFs "Listagem de Pontos" do GoolSystem (CIMA.pdf, REAL.pdf, SFRA.pdf)
e gera um arquivo Excel por empresa, com uma aba por linha (Atendimento
Principal = Sim E Ativo = Sim), contendo a listagem de pontos em ordem
(Nome, Nome Abrev., Endereco, Ordem, Vel. Limite, Latitude, Longitude).

Uso:
    cd goolsystem
    python gerar_excel_linhas.py [CIMA|REAL|SFRA ...]   (default: os 3)
"""
import re
import sys
from pathlib import Path

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

BASE_DIR = Path(__file__).resolve().parent

# limites x (em pontos PDF) de cada coluna da tabela de pontos, calibrados
# a partir das posicoes reais das palavras no cabecalho/primeira linha
COL_BOUNDS = [
    ("nome", 0, 95),
    ("nome_abrev", 95, 205),
    ("endereco", 205, 585),
    ("ordem", 585, 640),
    ("vel_limite", 640, 700),
    ("latitude", 700, 765),
    ("longitude", 765, 900),
]

LABELS_METADADO = ("Atendimento Principal:", "Nome Ida:", "Ativo:", "Nome Volta:")


def coluna_para_x(x0: float) -> str:
    for nome, ini, fim in COL_BOUNDS:
        if ini <= x0 < fim:
            return nome
    return "endereco"


def agrupar_por_linha_visual(words, tolerancia=2.5):
    """Agrupa palavras em 'linhas visuais' pela coordenada 'top' (y)."""
    linhas = []
    atual = []
    top_atual = None
    for w in sorted(words, key=lambda w: (w["top"], w["x0"])):
        if top_atual is None or abs(w["top"] - top_atual) <= tolerancia:
            atual.append(w)
            top_atual = w["top"] if top_atual is None else top_atual
        else:
            linhas.append(atual)
            atual = [w]
            top_atual = w["top"]
    if atual:
        linhas.append(atual)
    return linhas


def linha_e_metadado(texto: str) -> bool:
    return any(texto.startswith(lbl) or lbl in texto[:40] for lbl in LABELS_METADADO) or texto.startswith("Nome Nome Abrev")


def parsear_metadado(texto: str, estado: dict) -> bool:
    """Atualiza 'estado' com o campo encontrado nesta linha visual. Retorna
    True quando o campo era 'Nome Volta:' -- ultimo campo do cabecalho, ponto
    seguro pra decidir se um bloco novo comeca (nesse momento linha/nome_ida/
    ativo/principal ja foram todos atualizados pra o bloco atual)."""
    m = re.search(r"Atendimento Principal:\s*(Sim|N.o).*?Linha:\s*(.+)", texto)
    if m:
        estado["principal"] = m.group(1)
        estado["linha"] = m.group(2).strip()
        return False
    m = re.search(r"Nome Ida:\s*(.+)", texto)
    if m:
        estado["nome_ida"] = m.group(1).strip()
        return False
    m = re.search(r"Ativo:\s*(Sim|N.o)", texto)
    if m:
        estado["ativo"] = m.group(1)
        return False
    m = re.search(r"Nome Volta:\s*(.+)", texto)
    if m:
        estado["nome_volta"] = m.group(1).strip()
        return True
    return False


def parsear_ponto(grupo_palavras):
    """Recebe um grupo de palavras (uma linha visual) e tenta extrair um
    registro de ponto. Retorna None se a linha nao tiver os 4 campos
    numericos finais (ordem, vel, lat, lon) — nesse caso e provavelmente
    uma continuacao do endereco da linha anterior."""
    cols = {"nome": [], "nome_abrev": [], "endereco": [], "ordem": [], "vel_limite": [], "latitude": [], "longitude": []}
    for w in sorted(grupo_palavras, key=lambda w: w["x0"]):
        cols[coluna_para_x(w["x0"])].append(w["text"])

    ordem = " ".join(cols["ordem"])
    vel = " ".join(cols["vel_limite"])
    lat = " ".join(cols["latitude"])
    lon = " ".join(cols["longitude"])

    if not (re.fullmatch(r"\d+", ordem) and re.fullmatch(r"\d+", vel)
            and re.fullmatch(r"-?\d+,\d+", lat) and re.fullmatch(r"-?\d+,\d+", lon)):
        return None, " ".join(cols["nome"] + cols["nome_abrev"] + cols["endereco"])

    return {
        "nome": " ".join(cols["nome"]),
        "nome_abrev": " ".join(cols["nome_abrev"]),
        "endereco": " ".join(cols["endereco"]),
        "ordem": int(ordem),
        "vel_limite": int(vel),
        "latitude": lat.replace(",", "."),
        "longitude": lon.replace(",", "."),
    }, None


def extrair_linhas_do_pdf(pdf_path: Path) -> list[dict]:
    """Retorna lista de dicts {linha, nome_ida, nome_volta, pontos:[...]},
    somente para blocos com Atendimento Principal = Sim e Ativo = Sim."""
    resultado = []
    estado = {"principal": None, "linha": None, "nome_ida": None, "ativo": None, "nome_volta": None}
    bloco_atual = None  # dict em construcao, ligado por identidade a resultado[-1]

    with pdfplumber.open(str(pdf_path)) as pdf:
        total = len(pdf.pages)
        for i, page in enumerate(pdf.pages):
            if (i + 1) % 50 == 0:
                print(f"    ... pagina {i+1}/{total}")
            words = page.extract_words()
            if not words:
                continue
            for grupo in agrupar_por_linha_visual(words):
                texto = " ".join(w["text"] for w in sorted(grupo, key=lambda w: w["x0"]))

                if linha_e_metadado(texto):
                    cabecalho_completo = parsear_metadado(texto, estado)
                    if cabecalho_completo and estado["linha"]:
                        ativo_ok = (estado["ativo"] or "").lower().startswith("sim")
                        principal_ok = (estado["principal"] or "").lower().startswith("sim")
                        if ativo_ok and principal_ok:
                            bloco_atual = {
                                "linha": estado["linha"],
                                "nome_ida": estado["nome_ida"],
                                "nome_volta": estado["nome_volta"],
                                "pontos": [],
                            }
                            resultado.append(bloco_atual)
                        else:
                            bloco_atual = None
                    continue

                ponto, continuacao_endereco = parsear_ponto(grupo)
                if bloco_atual is None:
                    continue
                if ponto:
                    bloco_atual["pontos"].append(ponto)
                elif continuacao_endereco and bloco_atual["pontos"]:
                    bloco_atual["pontos"][-1]["endereco"] += " " + continuacao_endereco

    return resultado


def nome_aba_valido(nome: str, usados: set) -> str:
    limpo = re.sub(r'[\\/*?:\[\]]', "-", nome).strip()
    limpo = limpo[:31] if limpo else "linha"
    base = limpo
    n = 2
    while limpo in usados:
        sufixo = f" ({n})"
        limpo = base[: 31 - len(sufixo)] + sufixo
        n += 1
    usados.add(limpo)
    return limpo


def gerar_excel(empresa: str, linhas: list[dict], saida: Path):
    wb = Workbook()
    wb.remove(wb.active)
    usados = set()

    for bloco in linhas:
        aba_nome = nome_aba_valido(bloco["linha"], usados)
        ws = wb.create_sheet(title=aba_nome)

        ws["A1"] = "Linha:"
        ws["B1"] = bloco["linha"]
        ws["A2"] = "Nome Ida:"
        ws["B2"] = bloco["nome_ida"] or ""
        ws["A3"] = "Nome Volta:"
        ws["B3"] = bloco["nome_volta"] or ""
        for r in (1, 2, 3):
            ws.cell(row=r, column=1).font = Font(bold=True)

        cabecalho = ["Ordem", "Nome", "Nome Abrev.", "Endereço", "Vel. Limite", "Latitude", "Longitude"]
        header_row = 5
        for c, titulo in enumerate(cabecalho, start=1):
            cel = ws.cell(row=header_row, column=c, value=titulo)
            cel.font = Font(bold=True)
            cel.alignment = Alignment(horizontal="center")

        for i, p in enumerate(sorted(bloco["pontos"], key=lambda p: p["ordem"]), start=1):
            row = header_row + i
            ws.cell(row=row, column=1, value=p["ordem"])
            ws.cell(row=row, column=2, value=p["nome"])
            ws.cell(row=row, column=3, value=p["nome_abrev"])
            ws.cell(row=row, column=4, value=p["endereco"])
            ws.cell(row=row, column=5, value=p["vel_limite"])
            ws.cell(row=row, column=6, value=float(p["latitude"]))
            ws.cell(row=row, column=7, value=float(p["longitude"]))

        larguras = [8, 30, 16, 55, 11, 12, 12]
        for c, larg in enumerate(larguras, start=1):
            ws.column_dimensions[get_column_letter(c)].width = larg
        ws.freeze_panes = f"A{header_row + 1}"

    wb.save(saida)
    total_pontos = sum(len(b["pontos"]) for b in linhas)
    print(f"  [OK] {saida.name}  ({len(linhas)} linhas, {total_pontos} pontos)")


def main():
    empresas = sys.argv[1:] or ["CIMA", "REAL", "SFRA"]
    for empresa in empresas:
        pdf_path = BASE_DIR / f"{empresa}.pdf"
        if not pdf_path.exists():
            print(f"[AVISO] {pdf_path} nao encontrado, pulando.")
            continue
        print(f"Processando {empresa}.pdf ...")
        linhas = extrair_linhas_do_pdf(pdf_path)
        saida = BASE_DIR / f"{empresa}.xlsx"
        gerar_excel(empresa, linhas, saida)


if __name__ == "__main__":
    main()
