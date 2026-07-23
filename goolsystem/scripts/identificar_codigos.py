"""
Cruza as abas dos Excel gerados (gerar_excel_linhas.py) com o OSO mais
recente pra descobrir o codigo DMTT de cada linha, por similaridade de nome
DENTRO da mesma empresa (nome sozinho nao basta -- linhas de empresas
diferentes podem ter nomes parecidos).

Atualiza os .xlsx in-place: acrescenta "Codigo DMTT:" no cabecalho de cada
aba e prefixa o nome da aba com o codigo (quando encontrado).

Uso:
    cd goolsystem
    python identificar_codigos.py
"""
import re
import sys
import unicodedata
from copy import copy
from pathlib import Path

import openpyxl
from openpyxl.styles import Font

BASE_DIR = Path(__file__).resolve().parent
ROOT = BASE_DIR.parent
sys.path.insert(0, str(ROOT / "python"))
from comparar_oso_dados_unificados import extrair_oso  # noqa: E402

OSO_PDF = ROOT / "linhas" / "resumo de oso 08-07-2026.pdf"

EMPRESA_OSO = {
    "CIMA": "CIMA - VIAÇÃO CIDADE DE MACEIÓ LTDA",
    "REAL": "REAL - REAL TRANSPORTES URBANOS LTDA.",
    "SFRA": "SFRA - EMPRESA SÃO FRANCISCO LTDA.",
}

LIMIAR_SIMILARIDADE = 0.25


def normalizar_texto(t: str) -> str:
    t = unicodedata.normalize("NFKD", t or "")
    t = "".join(c for c in t if not unicodedata.combining(c))
    t = t.upper()
    t = re.sub(r"[^A-Z0-9]+", " ", t)
    return t.strip()


def palavras(t: str) -> set:
    return {w for w in normalizar_texto(t).split() if len(w) >= 3}


def similaridade(a: str, b: str) -> float:
    pa, pb = palavras(a), palavras(b)
    if not pa or not pb:
        return 0.0
    return len(pa & pb) / len(pa | pb)


def melhor_match(nomes_excel: list[str], candidatos: list[dict]) -> tuple[dict | None, float]:
    melhor, melhor_score = None, 0.0
    for cand in candidatos:
        score = max(similaridade(n, cand["nome"]) for n in nomes_excel if n)
        if score > melhor_score:
            melhor, melhor_score = cand, score
    return (melhor, melhor_score) if melhor_score >= LIMIAR_SIMILARIDADE else (None, melhor_score)


def nome_aba_valido(nome: str, usados: set) -> str:
    limpo = re.sub(r'[\\/*?:\[\]]', "-", nome).strip()[:31] or "linha"
    base = limpo
    n = 2
    while limpo in usados:
        sufixo = f" ({n})"
        limpo = base[: 31 - len(sufixo)] + sufixo
        n += 1
    usados.add(limpo)
    return limpo


def main():
    print("Extraindo OSO...")
    oso = extrair_oso(OSO_PDF)

    empresas_alvo = sys.argv[1:] or ["CIMA", "REAL", "SFRA"]

    for empresa in empresas_alvo:
        if empresa not in EMPRESA_OSO:
            print(f"[AVISO] empresa desconhecida: {empresa}, pulando.")
            continue
        arquivo_xlsx = f"{empresa}.xlsx"
        path = BASE_DIR / arquivo_xlsx
        if not path.exists():
            print(f"[AVISO] {path} nao encontrado, pulando.")
            continue

        wb_check = openpyxl.load_workbook(path, read_only=True)
        primeira_aba = wb_check[wb_check.sheetnames[0]]
        ja_processado = (primeira_aba["A1"].value or "") == "Código DMTT:"
        wb_check.close()
        if ja_processado:
            print(f"[PULANDO] {arquivo_xlsx} ja foi identificado antes (evita duplicar/corromper).")
            continue

        candidatos = [l for l in oso if l["empresa"] == EMPRESA_OSO[empresa] and l["tipo"] != "CATRACA DE SOLO"]
        print(f"\n=== {empresa} ({len(candidatos)} linhas ativas no OSO) ===")

        wb = openpyxl.load_workbook(path)
        usados_abas = set()
        encontrados, nao_encontrados = [], []

        for aba_nome_original in list(wb.sheetnames):
            ws = wb[aba_nome_original]
            linha_txt = ws["B1"].value or ""
            nome_ida_txt = ws["B2"].value or ""
            nome_volta_txt = ws["B3"].value or ""

            match, score = melhor_match([linha_txt, nome_ida_txt, nome_volta_txt], candidatos)

            ws.insert_rows(1)
            ws["A1"] = "Código DMTT:"
            ws["A1"].font = copy(ws["A2"].font)
            if match:
                ws["B1"] = f"{match['codigo']}  (similaridade {score:.0%}, OSO {match['oso']})"
                encontrados.append((match["codigo"], linha_txt))
                novo_nome = nome_aba_valido(f"{match['codigo']} - {linha_txt}", usados_abas)
            else:
                ws["B1"] = f"NAO IDENTIFICADO (melhor score {score:.0%})"
                nao_encontrados.append(linha_txt)
                novo_nome = nome_aba_valido(aba_nome_original, usados_abas)

            ws.title = novo_nome

        wb.save(path)
        print(f"  Identificadas: {len(encontrados)}/{len(encontrados) + len(nao_encontrados)}")
        if nao_encontrados:
            print("  NAO identificadas:")
            for n in nao_encontrados:
                print(f"    - {n}")


if __name__ == "__main__":
    main()
